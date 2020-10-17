from numpy import *
from moment import *
from rmatr import *
from const import *
from amatr import *
from eqn import *
from ddist import *
from bde import *
from dde import *
from solve2 import *
from wdist import *
from bee import *
from dee import *
from solve import *
from msdist import *
from msd3 import *
from findef2 import *
from findef import *
from tmoment import *
import sys

LIM1 = 40           #Limitation on 'No. of Harmonics'
LIM2 = 7            #Limitation on 'No. of Diaphragms'
LIM3 = 30           #Limitation on 'No. of Girders and Loads'
LIM4 = 30           #Limitation on 'No. of Reference Points'

FF = zeros(LIM4 + 1)
KGIR = zeros(LIM4 + 1)
XCOL = zeros(LIM4 + 1)

print("START")
print("READING DATA FROM INPUT FILE")

file = open("SECAN.DAT","r")
from openpyxl import *
ofile = load_workbook('D:\VNIT\Project\Phase\Programs\SECAN\SECAN\SECAN_output.xlsx')
ws = ofile.active
NB = int(file.readline())
row = 0
for I in range(1,NB+1):
    KZ = int(file.readline())
    TITLE = file.readline()
    N = int(file.readline())
    NG = int(file.readline())
    SPAN = float(file.readline())
    E = float(file.readline())
    G = float(file.readline())
    ND = int(file.readline())
    NCOL = int(file.readline())

    GS = zeros(LIM3 + 1)
    line = file.readline()
    i = 0
    for value in line.split():
        i += 1
        GS[i] = float(value)

    GMI = zeros(NG+1)
    line = file.readline()
    i = 0
    for value in line.split():
        i += 1
        GMI[i] = float(value)

    GTI = zeros(NG+1)
    line = file.readline()
    i = 0
    for value in line.split():
        i += 1
        GTI[i] = float(value)

    T = float(file.readline())
    EC = float(file.readline())
    GC = float(file.readline())
    SA = float(file.readline())

    if ND != 0:
        ED = float(file.readline())

        DLSD = zeros(ND+1)
        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            DLSD[i] = float(value)

        GMID = zeros(ND+1)
        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            GMID[i] = float(value)

    if NCOL > 0:
        DELTA = zeros(NCOL+1)
        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            DELTA[i] = float(value)

        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            FF[i] = float(value)

        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            KGIR[i] = float(value)

        line = file.readline()
        i = 0
        for value in line.split():
            i += 1
            XCOL[i] = float(value)

    NL = int(file.readline())

    #M,NW,NREF = zeros(NL+1),zeros(NL+1),zeros(NL+1)
    W,DLS,DLG,XREF = zeros(21),zeros(21),zeros(21),zeros(21),
    for i in range(1,NL+1):
        M = int(float(file.readline()))

        line = file.readline()
        j = 0
        for value in line.split():
            j += 1
            W[j] = float(value)

        line = file.readline()
        j = 0
        for value in line.split():
            j += 1
            DLS[j] = float(value)

        NW = int(float(file.readline()))

        line = file.readline()
        j = 0
        for value in line.split():
            j += 1
            DLG[j] = float(value)

        NREF = int(float(file.readline()))

        line = file.readline()
        j = 0
        for value in line.split():
            j += 1
            XREF[j] = float(value) 

        print("FINISHED READING DATA")
        print("START COMPUTING")

        if i == 1:
            if I == 1:
                ws['C1'] = NB
        row = row*(I-1) + 1
        ws.cell(row+1, 3).value = I
        ws.cell(row+2, 3).value = TITLE
        row = row+2
        KONT = 1

        if NB > 5: KONT += 1
        if N > LIM1: KONT += 1
        if NG > LIM3: KONT += 1
        if ND > LIM2: KONT += 1
        if NL > 5: KONT += 1
        if M > LIM3: KONT += 1
        if NW > LIM3: KONT += 1
        if NREF > LIM3: KONT += 1
        if NCOL > LIM4: KONT += 1
        
        if KONT > 1:
            sys.exit("Bridge input are out of limits")
        else:
            WF = zeros((21,11))

            NUMB = zeros(NREF+1)
            for j in range(1,NREF+1):
                NUMB[j] = j
        
            DG = zeros(NG+1)
            DG[1] = GS[1]
            for j in range(2,NG):
                DG[j] = GS[j] + DG[j-1]

            XT = zeros(NG+1)
            #XT[1] = 0
            for j in range(2,NG+1):
                XT[j] = DG[j-1]

            if i == 1:
                ws.cell(row+1, 3).value = N
                ws.cell(row+2, 3).value = NG
                ws.cell(row+3, 3).value = SPAN
                ws.cell(row+5, 3).value = E
                ws.cell(row+6, 3).value = G
                ws.cell(row+7, 3).value = NCOL
                row += 7 + 1    #row 11
                for j in range(1,NG+1):
                    ws.cell(row + j, 1).value = j
                    ws.cell(row + j, 2).value = GMI[j]
                    ws.cell(row + j, 3).value = GTI[j]
                row += j + 2    #row 16
                for j in range(1,NG):
                    ws.cell(row + j, 1).value = j
                    ws.cell(row + j, 2).value = GS[j]
                    ws.cell(row + j, 3).value = DG[j]
                row += j + 4    #row 22
                ws.cell(row, 1).value = T
                ws.cell(row, 2).value = EC
                ws.cell(row, 3).value = GC
                ws.cell(row, 4).value = SA
                ws.cell(row+2, 3).value = NL
                row += 4        #row 26
            ws.cell(row, 3).value = i
            ws.cell(row + 2, 3).value = M
            row += 3            #row 29

            for j in range(1,M+1):
                ws.cell(row + j, 1).value = j
                ws.cell(row + j, 2).value = W[j]
                ws.cell(row + j, 3).value = DLS[j]
            row += j + 2        #row 38
            ws.cell(row, 3).value = NW
            row += 1        #row 39
            for j in range(1, NW + 1):
                ws.cell(row + j, 1).value = j
                ws.cell(row + j, 3).value = DLG[j]
            row += j + 2    #row 43

            if i == 1:
                ws.cell(row, 3).value = NREF
                for j in range(1,NREF+1):
                    ws.cell(row+1, j+2).value = NUMB[j]
                #for j in range(1,NREF+1):
                    ws.cell(row+2, j+2).value = XREF[j]
                row += 4    #row 47
                if ND != 0:
                    ws.cell(row, 3).value = ND
                    for j in range(1,ND+1):
                        ws.cell(row+1, j+2).value = DLSD[j]
                    #for j in range(1,ND+1):
                        ws.cell(row+2, j+2).value = GMID[j]
                row += 4    #row 51
                if NCOL == 0:
                    FF[1] = 0
                    KGIR[1] = 1
                    XCOL[1] = 0
                    NCOL = 1
                if XCOL[1] != 0:
                    for j in range(1,NCOL+1):
                        #ofile.write("COLUMN DETAILS     PRESCRIBED")    ##Format in excel
                        ofile.write(str(DELTA[j]))
                    for j in range(1,NCOL+1):
                        ofile.write(str(FF[j]))
                    for j in range(1,NCOL+1):
                        ofile.write(str(KGIR[j]))
                    for j in range(1,NCOL+1):
                        ofile.write(str(XCOL[j]))
            SIG = zeros(LIM3 + 1)
            if SA != 0:
                for j in range(1,NG+1):
                    if j != NG:
                        B = GS[j]
                    SIG[j] = (0.5*EC*T**3)/(GC*SA*B*B)
            BM = zeros(LIM1 + 1)
            BM, row = MOMENT(M, N, W, DLS, SPAN, BM, KZ, row)
            if 1==1:
                #ws.cell(row, 1).value = "MOMENT COEFFICIENTS DUE TO ONE LINE OF WHEEL"
                #ws.cell(row+1, 1).value = "Harmonics"
                #ws.cell(row+1, 3).value = "Moment Coefficient"
                #for j in range(1, N + 1):
                    #ws.cell(row+1+j, 1).value = j
                    #ws.cell(row+1+j, 3).value = BM[j]
                #row += 1+j + 2  #row 59 
                pass    #To minimise. Remove later
            RM = zeros(LIM3*2 + 1)
            RM, row = RMATR(NG, NW, RM, DLG, DG, GS, KZ, SIG, row)
            if 1==1:         
                #ws.cell(row, 1).value = "CALCULATED R VECTOR"
                #ws.cell(row+1, 1).value = "Row No."
                #ws.cell(row+1,3).value = "Term"
                #for j in range(1, NG2 + 1):
                #    ws.cell(row+1+j, 1).value = j
                #    ws.cell(row+1+j, 3).value = RM[j]
                #row += 1+j + 2  #row 68
                pass           #To minimise. Remove later
            I1 = 30
            KT, MLC, C, LAMDA, ALFA, MU = zeros(LIM3 + 1), zeros(LIM3 + 1), zeros(LIM3 + 1), zeros(LIM3 + 1), zeros(LIM3 + 1), zeros(LIM3 + 1)
            KT, MLC, C, LAMDA, ALFA, MU, row = CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ, row)
            if 1==1:
                #ws.cell(row, 1).value = "CALCULATED CONSTANTS FOR HARMONIC NO."
                #ws.cell(row, 3).value = I1
                #ws.cell(row+1, 1).value = "Panel No."
                #ws.cell(row+1,3).value = "K"
                #ws.cell(row+1,4).value = "MLC"
                #ws.cell(row+1,5).value = "C"
                #ws.cell(row+1,6).value = "LAMDA"
                #ws.cell(row+1,7).value = "ALFA"
                #ws.cell(row+1,8).value = "MU"
                #for j in range(1, NG + 1):
                #    ws.cell(row+1 + j,1).value = j
                #    ws.cell(row+1 + j,3).value = KT[j]
                #    ws.cell(row+1 + j,4).value = MLC[j]
                #    ws.cell(row+1 + j,5).value = C[j]
                #    ws.cell(row+1 + j,6).value = LAMDA[j]
                #    ws.cell(row+1 + j,7).value = ALFA[j]
                #    ws.cell(row+1 + j,8).value = MU[j]
                #row += 1+j + 2  #row 74
                pass        #To minimise. Remove later
            AM = zeros((LIM3*2 + 1, LIM3*2 + 1))
            SIG,XYZ1,XYZ2,XYZ3,GN1,GN2,GN3,AM,row = AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG, row)
            if 1 == 1 :
                #if KZ == 2:
                #    ws.cell(row,1).value = "ND"
                #    ws.cell(row,3).value = "SIG[ND]"
                #    ws.cell(row,4).value = "XYZ3"
                #    for ND in range(1, NGL + 1):
                #        ws.cell(row + ND,1).value = ND
                #        ws.cell(row + ND,3).value = SIG[ND]
                #        ws.cell(row + ND,4).value = XYZ3
                #    row += ND + 2 #row 78
                #    ws.cell(row ,1).value = "ND"
                #    ws.cell(row ,3).value = "NE"
                #    ws.cell(row ,4).value = "SIG[ND]"
                #    ws.cell(row ,5).value = "XYZ1"
                #    ws.cell(row ,6).value = "XYZ2"
                #    ws.cell(row ,7).value = "GN1"
                #    ws.cell(row ,8).value = "GN2"
                #    ws.cell(row ,9).value = "GN3"
                #    for ND in range(1, NGL + 1):
                #        for NE in range(2, ND + 1):
                #            ws.cell(row + NE-1,1).value = ND
                #            ws.cell(row + NE-1,3).value = NE
                #            ws.cell(row + NE-1,4).value = SIG[ND]
                #            ws.cell(row + NE-1,5).value = XYZ1
                #            ws.cell(row + NE-1,6).value = XYZ2
                #            ws.cell(row + NE-1,7).value = GN1
                #            ws.cell(row + NE-1,8).value = GN2
                #            ws.cell(row + NE-1,9).value = GN3
                #row += (NE-1) + 2 #row 81
                #if KZ != 1:
                #    ws.cell(row, 1).value = "CALCULATED AMATRIX FOR HARMONICS NO:"
                #    ws.cell(row, 3).value = I1
                #    for j in range(1, NG2 + 1):
                #        for k in range(1, NG2 + 1):
                #            ws.cell(row+j,k).value = AM[j][k]
                #row += j + 2 #row 89
                #ofile.save('SECAN_output.xlsx')
                pass        #To minimise. Remove later
            ICNK = 1
            XX = zeros(LIM3*2 + 1)
            HARC = zeros((LIM1 + 1, LIM3*2 + 1))
            XX, HARC, row = EQN(I1, NG, AM, RM, XX, HARC, KZ, ICNK, row)
            if 1==1:
                #if KZ != 1:
                #    ws.cell(row, 1).value = "CORRELATION COEFFICIENTS"
                #    for j in range(1, KI + 1):
                #        ws.cell(row + j, 1).value = "B" + str(j)
                #        ws.cell(row + j, 3).value = XX[j]
                #ofile.save('SECAN_output.xlsx')
                #row = row + j + 2   #row 97
                pass      #To minimise. Remove later  
            ICNK = 2

            #CINF = zeros((NG + 1, 1 + 1))
            CINF = zeros((LIM3 + 1, LIM3*LIM2 + 1))
            for j in range(1,NG+1):
                CINF[j][1] = XX[j]
        
            CCINF = zeros((NG*2 + 1, N + 1))
            for I1 in range(1,N+1):
                KT, MLC, C, LAMDA, ALFA, MU, row = CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ, row)
                SIG,XYZ1,XYZ2,XYZ3,GN1,GN2,GN3,AM,row = AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG, row)
                XX, HARC, row = EQN(I1, NG, AM, RM, XX, HARC, KZ, ICNK, row)
                for j in range(1,NG*2 + 1):
                    CCINF[j][I1] = XX[j]

            if ND != 0:
                LCONT = 1
                WD = zeros(LIM2*LIM3 + 1)
                WDD = zeros((LIM2*LIM3, LIM2*LIM3 + 1))
                I6 = 0
                DDIST(N, NG, ND, DLSD, WD, M, W, DLS, CINF, GMI, E, SPAN, HARC, BM, LCONT, WDD, I6, KZ)
                DLGD = zeros(LIM3+1)
                NWI = 1
                for j in range(1, ND + 1):
                    for k in range(1, NG + 1):
                        I6 = k + NG*(j - 1)
                        I4 = k
                        if I4 < NG: DLGD[1] = DG[I4] - GS[I4]
                        if I4 == NG: DLGD[1] = DG[I4 - 1]
                        RM, row = RMATR(NG, NWI, RM, DLGD, DG, GS, KZ, SIG, row)
                        NPLUS = N + 1
                        for I7 in range(1, NPLUS + 1):
                            if I7 > 1:
                                I1 = I7 - 1
                                ICNK = 2
                            else:
                                I1 = 30
                                ICNK = 1
                            KT, MLC, C, LAMDA, ALFA, MU, row = CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ, row)
                            SIG,XYZ1,XYZ2,XYZ3,GN1,GN2,GN3,AM,row = AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG, row)
                            HARC1 = zeros((LIM1 + 1, LIM3 + 1))
                            XX, HARC1, row = EQN(I1, NG, AM, RM, XX, HARC1, KZ, ICNK, row)
                            CINFD = zeros((LIM3 + 1,LIM2*LIM3+1 + 1))
                            if I7 == 1:
                                I9 = I6 + 1
                                for I8 in range(1, NG + 1):
                                    CINFD[I8][I9] = XX[I8]
                        W1, DLS1 = zeros(LIM3 + 1), zeros(LIM3 + 1)
                        M1 = 1
                        W1[1] = 1
                        DLS1[1] = DLSD[j]
                        BM2 = zeros(LIM1 + 1)
                        BM2, row = MOMENT(M1, N, W1, DLS1, SPAN, BM2, KZ, row)
                        LCONT = 2
                        DDIST(N ,NG, ND, DLSD, WD, M1, W1, DLS1, CINFD, GMI, E, SPAN, HARC1, BM2, LCONT, WDD, I6, KZ)

            #BDE(ND, NG, ED, GS, DG, GMID, WDD, BID)
            #DDE(DG, WD, DID, NG, ND)
            #SOLVE2(NG, ND, BID, DID)
            for MM in range(1, NG*ND + 1):
                ARED[MM] = -DID[MM]

            if XCOL[1] != 0:
                LCONT = 1
                #WDIST(N, NCOL, KGIR, XCOL, WB, M, W, DLS, CINF, GMI, E, SPAN, HARC, BM, LCONT, WU, I3, KZ)
                for I1 in range(1, LIM3 + 1):
                    DLGI[I1] = 0

                NWI = 1
                for I3 in range(1, NCOL+1):
                    I4 = KGIR[I3]
                    if I4 < NG: DLGI[1] = DG[I4] - GS[I4]
                    if I4 == NG: DLGI[1] = DG[I4-1]
                    #RMATR(NG, NWI, RM, DLGI, DG, GS, KZ, SIG)
                    NPLUS = N + 1
                    for I7 in range(1, NPLUS+1):
                        if I7 > 1:
                            I1 = I7 - 1
                            ICNK = 2
                        else:
                            I1 = 30
                            ICNK = 1
                        #CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ)
                        #AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG)
                        #EQN(I1, NG, AM, RM, XX, HARC1, KZ, ICNK)
                        if I7 <= 1:
                            I9 = I3 + 1
                            for I8 in range(1, NG+1):
                                CINF[I8][I9] = XX[I8]
                    M1 = 1
                    W1[1] = 1
                    DLS1[1] = XCOL[I3]
                    #MOMENT(M1, N, W1, DLS1, SPAN, BM1, KZ)
                    LCONT = 2
                    #WDIST(N, NCOL, KGIR, XCOL, WB, M1, W1, DLS1, CINF, GM1, E, SPAN, HARC1, BM1, LCONT, WU, I3, KZ)

            #BEE(FF, WU, BI, NCOL)
            #DEE(WB, DELTA, DI, NCOL)
            #SOLVE(BI, NCOL, DI, XCOL)
       
            for MM in range(1, NCOL + 1):
                ARE[MM] = DI[MM]
            if XCOL[1] != 0: pass #pending

            #MSDIST(N, NG, SPAN, BM, HARC, ABM, NREF, XREF, M, W, DLS, AS, CINF, KZ, XCOL, ND)
            for j in range(1,ND+1):
                for k in range(1,NG+1):
                    I6 = k + (j - 1)*NG
                    I4 = k
                    I5 = j
                    if I4 < NG: DLGD[1] = DG[I4] - GS[I4]
                    if I4 == NG: DLGD[1] = DG[I4-1]
                    if DLSD[1] != 0:
                        #RMATR(NG, NWI, RM, DLGD, DG, GS, KZ, SIG)
                        for I1 in range(1, N + 1):
                            #CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ)
                            #AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG)
                            #EQN(I1, NG, AM, RM, XX, HARC1, KZ, ICNK)
                            pass #to be removed
                    M1 = 1
                    W1[1] = -ARED[I6]
                    DLSI[1] = DLSD[I5]
                    if DLSD[1] != 0:
                        #MOMENT(M1, N, W1, DLSI, SPAN, BM2, KZ)
                        #MSD3(I6, N, NG, SPAN, BM2, HARC1, ABM, NREF, XREF, M1, W1, DLSI, AS, CINFD, KZ, XCOL, ND)
                        pass #to be removed
                    #FINDEF2(I6, I4, I5, N ,NG, ARED, DLSD, CINFD, GMI, E, SPAN, HARC1, KZ, NREF, XREF, WF)
            for I3 in range(1, NCOL + 1):
                I4 = KGIR[I3]
                if I4 < NG: DLGI[1] = DG[I4] - GS[I4]
                if I4 == NG: DLGI[1] = DG[I4-1]
                if XCOL[1] != 0:
                    #RMATR(NG, NWI, RM, DLGI, DG, GS, KZ, SIG)
                    pass #to be removed
                    for I1 in range(1, N + 1):
                        #CONST(I1, NG, EC, GC, T, KT, MLC, C, LAMDA, ALFA, MU, GMI, GTI, SPAN, GS, G, E, KZ)
                        #AMATR(I1, NG, DG, GS, AM, KT, LAMDA, ALFA, MU, KZ, SIG)
                        #EQN(I1, NG, AM, RM, XX, HARC1, KZ, ICNK)
                        pass #to be removed
                M1 = 1
                W1[1] = -ARE[I3]
                DLSI[1] = XCOL[I3]
                if XCOL[1] != 0:
                    #MOMENT(M1, N, W1, DLSI, SPAN, BM1, KZ)
                    #MSD2(I3, N, NG, SPAN, BM1, HARC1, ABM, NREF, XREF, M1, W1, DLS1, AS, CINF, KZ, NCOL)
                    pass
                #FINDEF(I3, N, NCOL, NG, KGIR, XCOL, M, W, ARE, DLS, CINF, GMI, E, SPAN, HARC, HARC1, BM, KZ, NREF, XREF, WF)

























